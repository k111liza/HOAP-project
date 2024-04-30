import openpyxl
from openpyxl.reader.excel import load_workbook

from scripts.resource_path import resource_path

months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]


def create_rent(id, surname, name, month, year, cold_water, hot_water, electricity, price_cold_water, price_hot_water, price_electricity):
    month =months[month]
    wb = load_workbook(resource_path('rent/rent-template.xlsx'))
    ws = wb.active
    ws['Y13']=f'{surname} {name}'
    ws['F11'] = f'{month} {year}'
    ws['BH29'] = f'{cold_water}'
    ws['BH30'] = f'{hot_water}'
    ws['BH38'] = f'{electricity}'
    ws['EI29'] = f'{price_cold_water}'
    ws['EI30'] = f'{price_hot_water}'
    ws['EI38'] = f'{price_electricity}'
    ws['GD43'] =f'{round(price_electricity+price_hot_water+price_cold_water, 2)}'
    wb.save(resource_path(f'rent/{id}_{surname}_{name}_{month}_{year}.xlsx'))

# create_rent(1,"Иванов", "Иван", 0, 2024, 124, 125, 789, 1212, 4554, 7878)
